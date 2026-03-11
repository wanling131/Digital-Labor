from __future__ import annotations

import argparse
import datetime as dt
import os
from typing import Dict, List, Optional, Tuple

from sqlalchemy import Engine, text

try:
    from .common import get_database_url, get_engine
except ImportError:
    from common import get_database_url, get_engine


def _load_workbook(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd  # type: ignore

        return xlrd.open_workbook(path)
    else:
        import openpyxl  # type: ignore

        return openpyxl.load_workbook(path, data_only=True)


def _iter_rows(path: str):
    wb = _load_workbook(path)
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        sheet = wb.sheet_by_index(0)
        header = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows):
            yield header, [sheet.cell_value(r, c) for c in range(sheet.ncols)]
    else:
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        header = [str(h or "").strip() for h in rows[0]]
        for r in rows[1:]:
            yield header, list(r)


def _guess_indices(header: List[str]) -> Dict[str, int]:
    def find(patterns: List[str]) -> int:
        for i, h in enumerate(header):
            h_lower = h.lower()
            for p in patterns:
                if p in h_lower:
                    return i
        return -1

    return {
        "name": find(["姓名", "name"]),
        "work_no": find(["工号", "工人编号", "workno", "work_no"]),
        "date": find(["日期", "考勤日期", "workdate", "date"]),
        "hours": find(["工时", "时长", "小时", "hours"]),
    }


def _parse_date(raw) -> Optional[dt.date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, dt.datetime):
        return raw.date()
    if isinstance(raw, dt.date):
        return raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:  # noqa: BLE001
            continue
    # Excel 可能给出数字序号，这里简单忽略复杂情况
    return None


def _find_person_id(conn, name: Optional[str], work_no: Optional[str]) -> Optional[int]:
    if work_no:
        r = conn.execute(
            text("SELECT id FROM person WHERE work_no = :w LIMIT 1"),
            {"w": work_no},
        ).mappings().first()
        if r:
            return int(r["id"])
    if name:
        r = conn.execute(
            text("SELECT id FROM person WHERE name = :n LIMIT 1"),
            {"n": name},
        ).mappings().first()
        if r:
            return int(r["id"])
    return None


def import_attendance_from_file(engine: Engine, excel_path: str) -> int:
    if not os.path.isfile(excel_path):
        print(f"[attendance] 文件不存在，跳过：{excel_path}")
        return 0

    inserted = 0
    skipped_no_person = 0

    with engine.begin() as conn:
        header: Optional[List[str]] = None
        idx = {}

        for header_row, row in _iter_rows(excel_path):
            if header is None:
                header = header_row
                idx = _guess_indices(header)
                print(f"[attendance] 检测到表头：{header}")
                print(f"[attendance] 列映射：{idx}")

            def val(key: str) -> Optional[str]:
                i = idx.get(key, -1)
                if i < 0 or i >= len(row):
                    return None
                v = row[i]
                if v is None:
                    return None
                return str(v).strip()

            name = val("name")
            work_no = val("work_no")
            date_raw = None
            di = idx.get("date", -1)
            if di >= 0 and di < len(row):
                date_raw = row[di]
            work_date = _parse_date(date_raw)

            hours_val = None
            hi = idx.get("hours", -1)
            if hi >= 0 and hi < len(row):
                hours_val = row[hi]
            try:
                hours = float(hours_val) if hours_val not in (None, "") else 0.0
            except Exception:  # noqa: BLE001
                hours = 0.0

            if not work_date:
                continue

            person_id = _find_person_id(conn, name=name, work_no=work_no)
            if not person_id:
                skipped_no_person += 1
                continue

            # 若已存在同人同日记录，则更新工时；否则插入新记录
            existing = conn.execute(
                text(
                    """
                    SELECT id FROM attendance
                    WHERE person_id = :pid AND work_date = :d
                    """
                ),
                {"pid": person_id, "d": work_date},
            ).mappings().first()

            if existing:
                conn.execute(
                    text(
                        """
                        UPDATE attendance
                        SET hours = :h
                        WHERE id = :id
                        """
                    ),
                    {"h": hours, "id": int(existing["id"])},
                )
            else:
                conn.execute(
                    text(
                        """
                        INSERT INTO attendance (person_id, org_id, work_date, hours)
                        VALUES (:pid, NULL, :d, :h)
                        """
                    ),
                    {"pid": person_id, "d": work_date, "h": hours},
                )
                inserted += 1

    print(f"[attendance] 导入完成，新增/更新 {inserted} 条，跳过 {skipped_no_person} 条（未找到人员）")
    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(description="Import attendance from Excel into PostgreSQL attendance table.")
    parser.add_argument("--database-url", default="", help="Override DATABASE_URL for PostgreSQL.")
    parser.add_argument("--excel", required=True, help="Path to 考勤 Excel 文件（.xls/.xlsx）。")
    args = parser.parse_args()

    url = get_database_url(args.database_url or None)
    engine = get_engine(url)
    import_attendance_from_file(engine, args.excel)


if __name__ == "__main__":
    main()

