from __future__ import annotations

import argparse
import os
from typing import Any, Dict, List, Optional

from sqlalchemy import Engine, text

try:
    from .common import get_database_url, get_engine
except ImportError:
    from common import get_database_url, get_engine


def _load_workbook(path: str):
    """
    同时兼容 .xls 与 .xlsx：
    - .xlsx 使用 openpyxl
    - .xls 使用 xlrd
    """

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd  # type: ignore

        return xlrd.open_workbook(path)
    else:
        import openpyxl  # type: ignore

        return openpyxl.load_workbook(path, data_only=True)


def _guess_column_indices(header: List[str]) -> Dict[str, int]:
    """
    根据表头中文/英文关键词猜测各字段所在列。
    实际 Excel 表头可能略有出入，如有需要可按打印的 header 调整脚本。
    """

    def find(patterns: List[str]) -> int:
        for i, h in enumerate(header):
            h_lower = h.lower()
            for p in patterns:
                if p in h_lower:
                    return i
        return -1

    return {
        "name": find(["姓名", "name"]),
        "id_card": find(["身份证", "证件号", "idcard", "id_card"]),
        "mobile": find(["手机号", "手机", "联系电话", "mobile", "phone"]),
        "work_no": find(["工号", "工人编号", "workno", "work_no"]),
        "job_title": find(["工种", "岗位", "job", "job_title"]),
        "work_address": find(["项目", "工地", "地址", "work_address"]),
    }


def _iter_rows(path: str):
    wb = _load_workbook(path)
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        # xlrd
        sheet = wb.sheet_by_index(0)
        header = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows):
            yield header, [sheet.cell_value(r, c) for c in range(sheet.ncols)]
    else:
        # openpyxl
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        header = [str(h or "").strip() for h in rows[0]]
        for r in rows[1:]:
            yield header, list(r)


def import_persons(engine: Engine, excel_path: str) -> int:
    """
    从实名制 Excel 导入人员到 person 表。

    - 若已存在相同身份证或工号，则跳过该行。
    """

    if not os.path.isfile(excel_path):
        print(f"[person] 文件不存在，跳过：{excel_path}")
        return 0

    inserted = 0

    with engine.begin() as conn:
        header: Optional[List[str]] = None
        col_idx: Dict[str, int] = {}

        for header_row, row in _iter_rows(excel_path):
            if header is None:
                header = header_row
                col_idx = _guess_column_indices(header)
                print(f"[person] 检测到表头：{header}")
                print(f"[person] 列映射：{col_idx}")
            # 逐行插入
            name_idx = col_idx.get("name", -1)
            if name_idx < 0:
                continue
            raw_name = row[name_idx]
            name = str(raw_name or "").strip()
            if not name:
                continue

            def val(key: str) -> Optional[str]:
                idx = col_idx.get(key, -1)
                if idx < 0 or idx >= len(row):
                    return None
                v = row[idx]
                if v is None:
                    return None
                return str(v).strip()

            id_card = val("id_card")
            work_no = val("work_no")
            mobile = val("mobile")
            job_title = val("job_title")
            work_address = val("work_address")

            # 唯一性：优先身份证，其次工号
            exists = False
            if id_card:
                q = conn.execute(
                    text("SELECT 1 FROM person WHERE id_card = :id LIMIT 1"),
                    {"id": id_card},
                ).first()
                exists = bool(q)
            if not exists and work_no:
                q = conn.execute(
                    text("SELECT 1 FROM person WHERE work_no = :w LIMIT 1"),
                    {"w": work_no},
                ).first()
                exists = bool(q)
            if exists:
                continue

            conn.execute(
                text(
                    """
                    INSERT INTO person (org_id, work_no, name, id_card, mobile, status, contract_signed, on_site, job_title, work_address)
                    VALUES (:org_id, :work_no, :name, :id_card, :mobile, :status, :contract_signed, :on_site, :job_title, :work_address)
                    """
                ),
                {
                    "org_id": None,
                    "work_no": work_no,
                    "name": name,
                    "id_card": id_card,
                    "mobile": mobile,
                    "status": "预注册",
                    "contract_signed": 0,
                    "on_site": 0,
                    "job_title": job_title,
                    "work_address": work_address,
                },
            )
            inserted += 1

    print(f"[person] 导入完成，新增 {inserted} 条")
    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(description="Import persons from Excel into PostgreSQL person table.")
    parser.add_argument("--database-url", default="", help="Override DATABASE_URL for PostgreSQL.")
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to实名制 Excel 文件（.xls/.xlsx）。",
    )
    args = parser.parse_args()

    url = get_database_url(args.database_url or None)
    engine = get_engine(url)
    import_persons(engine, args.excel)


if __name__ == "__main__":
    main()

