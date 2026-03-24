from __future__ import annotations

import base64
import io
import os
import re
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from sqlalchemy import bindparam, text

from digital_labor.crypto_compat import encrypt, safe_decrypt_then_mask
from digital_labor.db import get_engine
from digital_labor.paths import get_paths
from digital_labor.utils.permission import get_org_tree_ids


from digital_labor.utils.cache import get_cache, set_cache

_JOB_TITLES_CACHE_TTL_SECONDS = 60


# 使用通用缓存工具
def _job_titles_cache_get(key: str) -> Optional[dict]:
    return get_cache(key)


def _job_titles_cache_set(key: str, data: dict, ttl: int = _JOB_TITLES_CACHE_TTL_SECONDS) -> None:
    set_cache(key, data, ttl)


@dataclass(frozen=True)
class Page:
    limit: int
    offset: int
    page: int
    page_size: int


def _save_signature_data_url(worker_id: int, data_url: str) -> Optional[str]:
    """
    兼容 Node 行为：activation 允许传 dataURL；后端落盘并在 DB 存相对 server/ 的路径。
    """
    if not data_url.startswith("data:image"):
        return data_url
    comma = data_url.find(",")
    if comma < 0:
        return None
    b64 = data_url[comma + 1 :]
    try:
        buf = base64.b64decode(b64)
    except Exception:  # noqa: BLE001
        return None
    paths = get_paths()
    dir_abs = paths.uploads_signatures
    os.makedirs(dir_abs, exist_ok=True)
    filename = f"person_{worker_id}_{int(time.time() * 1000)}.png"
    abs_path = os.path.join(dir_abs, filename)
    with open(abs_path, "wb") as f:
        f.write(buf)
    rel = os.path.relpath(abs_path, paths.server_root).replace("\\", "/")
    return rel


def archive_list(
    *,
    filters: Dict[str, Any],
    page: Page,
    actor_org_id: Optional[int] = None,
) -> dict:
    status = filters.get("status")
    org_id = filters.get("org_id")
    if actor_org_id is not None:
        org_id = actor_org_id
    on_site = filters.get("on_site")
    filled = filters.get("filled")
    contract_signed = filters.get("contract_signed")
    keyword = (filters.get("keyword") or "").strip()
    job_title = filters.get("job_title")

    where: List[str] = []
    params: Dict[str, Any] = {"limit": page.limit, "offset": page.offset}
    bindparams_list: List[bindparam] = []
    if status:
        where.append("p.status = :status")
        params["status"] = status
    if org_id:
        # 实现基于组织树的数据范围控制
        org_ids = get_org_tree_ids(int(org_id))
        if org_ids:
            where.append("p.org_id IN :org_ids")
            params["org_ids"] = org_ids
            bindparams_list.append(bindparam("org_ids", expanding=True))
    if on_site == "1":
        where.append("p.on_site = 1")
    if on_site == "0":
        where.append("p.on_site = 0")
    if filled == "1":
        where.append("TRIM(COALESCE(p.id_card,'')) != '' AND TRIM(COALESCE(p.mobile,'')) != ''")
    if filled == "0":
        where.append("(TRIM(COALESCE(p.id_card,'')) = '' OR TRIM(COALESCE(p.mobile,'')) = '')")
    if contract_signed == "1":
        where.append("p.contract_signed = 1")
    if contract_signed == "0":
        where.append("p.contract_signed = 0")
    if job_title:
        where.append("p.job_title = :job_title")
        params["job_title"] = job_title
    if keyword:
        where.append("(LOWER(p.name) LIKE LOWER(:kw) OR LOWER(p.work_no) LIKE LOWER(:kw) OR LOWER(o.name) LIKE LOWER(:kw))")
        params["kw"] = f"%{keyword}%"

    where_clause = " AND " + " AND ".join(where) if where else ""
    engine = get_engine()
    with engine.connect() as conn:
        total_query = text("SELECT COUNT(*) FROM person p LEFT JOIN org o ON p.org_id = o.id WHERE 1=1" + where_clause)
        if bindparams_list:
            total_query = total_query.bindparams(*bindparams_list)
        total = conn.execute(total_query, params).scalar_one()
        rows_query = text(
            """
            SELECT p.*, o.name as org_name
            FROM person p LEFT JOIN org o ON p.org_id = o.id
            WHERE 1=1"""
            + where_clause +
            " ORDER BY p.id DESC LIMIT :limit OFFSET :offset"
        )
        if bindparams_list:
            rows_query = rows_query.bindparams(*bindparams_list)
        rows = conn.execute(rows_query, params).mappings().all()

    out = []
    for r in rows:
        d = dict(r)
        if d.get("id_card"):
            d["id_card"] = safe_decrypt_then_mask(d["id_card"], "idCard")
        if d.get("mobile"):
            d["mobile"] = safe_decrypt_then_mask(d["mobile"], "mobile")
        if d.get("bank_card"):
            d["bank_card"] = safe_decrypt_then_mask(d["bank_card"], "bankCard")
        out.append(d)
    return {"list": out, "total": int(total), "page": page.page, "pageSize": page.page_size}


def archive_get(person_id: int) -> Optional[dict]:
    engine = get_engine()
    with engine.connect() as conn:
        row = conn.execute(
            text(
                """
                SELECT p.*, o.name as org_name
                FROM person p LEFT JOIN org o ON p.org_id=o.id
                WHERE p.id = :id
                """
            ),
            {"id": person_id},
        ).mappings().first()
    return dict(row) if row else None


def _validate_person_fields(body: Dict[str, Any], for_create: bool = True) -> None:
    """校验人员档案字段，防止无效或超长输入。"""
    if for_create and not (body.get("name") or "").strip():
        raise ValueError("姓名必填")
    if "name" in body:
        name = (body.get("name") or "").strip()
        if name and len(name) > 100:
            raise ValueError("姓名长度不能超过100字符")
    if "work_no" in body:
        work_no = (body.get("work_no") or "").strip()
        if work_no and len(work_no) > 50:
            raise ValueError("工号长度不能超过50字符")
    if "id_card" in body and body.get("id_card"):
        id_card = str(body["id_card"]).strip()
        if id_card and (len(id_card) < 15 or len(id_card) > 18):
            raise ValueError("身份证号长度应为15或18位")
    if "mobile" in body and body.get("mobile"):
        mobile = str(body["mobile"]).strip()
        if mobile and (len(mobile) != 11 or not mobile.isdigit()):
            raise ValueError("手机号应为11位数字")


def archive_create(body: Dict[str, Any]) -> int:
    _validate_person_fields(body, for_create=True)
    engine = get_engine()
    with engine.begin() as conn:
        params = {
            "org_id": body.get("org_id"),
            "work_no": body.get("work_no"),
            "name": body["name"],
            "id_card": encrypt(body["id_card"]) if body.get("id_card") else None,
            "mobile": encrypt(body["mobile"]) if body.get("mobile") else None,
            "bank_card": encrypt(body["bank_card"]) if body.get("bank_card") else None,
            "status": body.get("status") or "预注册",
            "job_title": body.get("job_title"),
        }
        if engine.dialect.name == "sqlite":
            rr = conn.execute(
                text(
                    """
                    INSERT INTO person (org_id, work_no, name, id_card, mobile, bank_card, status, job_title, updated_at)
                    VALUES (:org_id, :work_no, :name, :id_card, :mobile, :bank_card, :status, :job_title, CURRENT_TIMESTAMP)
                    """
                ),
                params,
            )
            return int(rr.lastrowid or 0)
        r = conn.execute(
            text(
                """
                INSERT INTO person (org_id, work_no, name, id_card, mobile, bank_card, status, job_title, updated_at)
                VALUES (:org_id, :work_no, :name, :id_card, :mobile, :bank_card, :status, :job_title, CURRENT_TIMESTAMP)
                RETURNING id
                """
            ),
            params,
        ).mappings().first()
        return int(r["id"])


def archive_update(person_id: int, patch: Dict[str, Any]) -> bool:
    _validate_person_fields(patch, for_create=False)
    engine = get_engine()
    with engine.connect() as conn:
        exists = conn.execute(text("SELECT 1 FROM person WHERE id = :id"), {"id": person_id}).first()
    if not exists:
        return False

    updates = ["updated_at = CURRENT_TIMESTAMP"]
    params: Dict[str, Any] = {"id": person_id}

    def set_field(col: str, key: str, transform=lambda x: x):  # noqa: B008
        if key in patch:
            updates.append(f"{col} = :{key}")
            params[key] = transform(patch.get(key))

    set_field("org_id", "org_id")
    set_field("work_no", "work_no")
    set_field("name", "name")
    set_field("id_card", "id_card", lambda v: encrypt(v) if v else None)
    set_field("mobile", "mobile", lambda v: encrypt(v) if v else None)
    set_field("bank_card", "bank_card", lambda v: encrypt(v) if v else None)
    set_field("status", "status")
    set_field("job_title", "job_title")
    set_field("work_address", "work_address")
    set_field("signature_image", "signature_image")

    if len(params.keys()) == 1:
        raise ValueError("无有效字段")

    with engine.begin() as conn:
        set_clause = ", ".join(updates)
        query = text(f"UPDATE person SET {set_clause} WHERE id = :id")
        conn.execute(query, params)
    return True


def me_activation(worker_id: int, patch: Dict[str, Any]) -> None:
    engine = get_engine()
    with engine.connect() as conn:
        exists = conn.execute(text("SELECT 1 FROM person WHERE id = :id"), {"id": worker_id}).first()
    if not exists:
        raise LookupError("人员不存在")

    updates = ["updated_at = CURRENT_TIMESTAMP"]
    params: Dict[str, Any] = {"id": worker_id}

    if "id_card" in patch:
        updates.append("id_card = :id_card")
        params["id_card"] = encrypt(str(patch.get("id_card"))) if patch.get("id_card") else None
    if "mobile" in patch:
        updates.append("mobile = :mobile")
        params["mobile"] = encrypt(str(patch.get("mobile"))) if patch.get("mobile") else None
    if "signature_image" in patch:
        sig = patch.get("signature_image")
        if sig in ("", None):
            stored = None
        else:
            stored = _save_signature_data_url(int(worker_id), str(sig))
        updates.append("signature_image = :signature_image")
        params["signature_image"] = stored

    if len(params.keys()) == 1:
        raise ValueError("无有效字段")

    with engine.begin() as conn:
        set_clause = ", ".join(updates)
        query = text(f"UPDATE person SET {set_clause} WHERE id = :id")
        conn.execute(query, params)


def batch_import_from_excel(data: bytes, *, default_job_title: Optional[str] = None, skip_errors: bool = True) -> dict:
    """
    批量导入人员，Excel 格式同 Node 端。工种为空时使用 default_job_title。

    Args:
        data: Excel 文件字节数据
        default_job_title: 默认工种
        skip_errors: 是否跳过错误行继续导入（默认 True）

    Returns:
        dict: {
            "imported": 导入成功数量,
            "total": 总行数,
            "errors": 错误列表,
            "error_rows": 错误行号列表（用于导出）,
            "progress": {"current": 当前行, "total": 总行数}
        }
    """
    try:
        import openpyxl  # type: ignore
    except Exception:  # noqa: BLE001
        raise RuntimeError("缺少依赖 openpyxl，无法解析 Excel")

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "total": 0, "errors": [], "error_rows": []}

    header = [str(h or "").strip() for h in rows[0]]

    def find_idx(*patterns: str) -> int:
        for i, h in enumerate(header):
            for p in patterns:
                if p.lower() in (h or "").lower():
                    return i
        return -1

    name_idx = find_idx("name", "姓名")
    work_no_idx = find_idx("work_no", "工号")
    id_card_idx = find_idx("id_card", "身份证号", "身份证")
    mobile_idx = find_idx("mobile", "手机号", "电话")
    org_idx = find_idx("org_id", "组织id", "所属组织")
    status_idx = find_idx("status", "状态")
    job_idx = find_idx("job_title", "工种", "job_type")

    if name_idx < 0:
        raise ValueError("Excel 需包含「姓名」列")

    mobile_re = re.compile(r"^1[3-9]\d{9}$")
    id_card_re = re.compile(r"^[1-9]\d{5}(18|19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])\d{3}[0-9Xx]$")

    errors: List[dict] = []
    error_rows: List[int] = []
    valid: List[Dict[str, Any]] = []
    total_rows = len(rows) - 1

    for i in range(1, len(rows)):
        r = rows[i]
        row_errors: List[str] = []
        name = str(r[name_idx] or "").strip() if name_idx >= 0 and r[name_idx] is not None else ""
        if not name:
            row_errors.append("姓名不能为空")

        mobile = ""
        if mobile_idx >= 0 and r[mobile_idx] is not None:
            mobile = str(r[mobile_idx]).strip()
        if mobile and not mobile_re.match(mobile):
            row_errors.append("手机号格式不正确")

        id_card = ""
        if id_card_idx >= 0 and r[id_card_idx] is not None:
            id_card = str(r[id_card_idx]).strip()
        if id_card and not id_card_re.match(id_card):
            row_errors.append("身份证号格式不正确")

        if row_errors:
            errors.append({"row": i + 2, "errors": row_errors, "data": {
                "name": name,
                "mobile": mobile,
                "id_card": id_card,
            }})
            error_rows.append(i + 2)
            continue

        work_no = str(r[work_no_idx]).strip() if work_no_idx >= 0 and r[work_no_idx] is not None else None
        work_no = work_no or None
        org_id = None
        if org_idx >= 0 and r[org_idx] is not None:
            try:
                org_id = int(str(r[org_idx]).strip())
            except Exception:  # noqa: BLE001
                pass
        status = str(r[status_idx]).strip() if status_idx >= 0 and r[status_idx] is not None else "预注册"
        status = status or "预注册"
        job_title = str(r[job_idx]).strip() if job_idx >= 0 and r[job_idx] is not None else ""
        job_title = job_title or default_job_title or None

        valid.append({
            "org_id": org_id,
            "work_no": work_no,
            "name": name,
            "id_card": id_card or None,
            "mobile": mobile or None,
            "bank_card": None,
            "status": status,
            "job_title": job_title,
        })

    # 如果不跳过错误且有错误，则不导入
    if not skip_errors and errors:
        return {
            "imported": 0,
            "total": total_rows,
            "errors": errors,
            "error_rows": error_rows,
            "progress": {"current": total_rows, "total": total_rows}
        }

    engine = get_engine()
    count = 0
    with engine.begin() as conn:
        for idx, item in enumerate(valid):
            params = {
                "org_id": item["org_id"],
                "work_no": item["work_no"],
                "name": item["name"],
                "id_card": encrypt(item["id_card"]) if item.get("id_card") else None,
                "mobile": encrypt(item["mobile"]) if item.get("mobile") else None,
                "bank_card": item.get("bank_card"),
                "status": item.get("status") or "预注册",
                "job_title": item.get("job_title"),
            }
            if engine.dialect.name == "sqlite":
                conn.execute(
                    text(
                        """
                        INSERT INTO person (org_id, work_no, name, id_card, mobile, bank_card, status, job_title, updated_at)
                        VALUES (:org_id, :work_no, :name, :id_card, :mobile, :bank_card, :status, :job_title, CURRENT_TIMESTAMP)
                        """
                    ),
                    params,
                )
            else:
                conn.execute(
                    text(
                        """
                        INSERT INTO person (org_id, work_no, name, id_card, mobile, bank_card, status, job_title, updated_at)
                        VALUES (:org_id, :work_no, :name, :id_card, :mobile, :bank_card, :status, :job_title, CURRENT_TIMESTAMP)
                        """
                    ),
                    params,
                )
            count += 1

    return {
        "imported": count,
        "total": total_rows,
        "errors": errors,
        "error_rows": error_rows,
        "progress": {"current": total_rows, "total": total_rows}
    }


def generate_error_excel(original_data: bytes, error_rows: List[int]) -> bytes:
    """
    生成错误行 Excel 文件。

    Args:
        original_data: 原始 Excel 文件字节数据
        error_rows: 错误行号列表（Excel 行号，从 1 开始）

    Returns:
        bytes: 错误行 Excel 文件字节数据
    """
    try:
        import openpyxl  # type: ignore
    except Exception:  # noqa: BLE001
        raise RuntimeError("缺少依赖 openpyxl，无法解析 Excel")

    wb = openpyxl.load_workbook(io.BytesIO(original_data), data_only=True)
    ws = wb.worksheets[0]

    # 创建新工作簿
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 复制标题行
    for col_idx, cell in enumerate(ws[1], 1):
        new_ws.cell(row=1, column=col_idx, value=cell.value)

    # 添加错误原因列
    error_col = ws.max_column + 1
    new_ws.cell(row=1, column=error_col, value="错误原因")

    # 复制错误行
    new_row_idx = 2
    for row_idx in error_rows:
        if row_idx <= ws.max_row:
            for col_idx, cell in enumerate(ws[row_idx], 1):
                new_ws.cell(row=new_row_idx, column=col_idx, value=cell.value)
            new_row_idx += 1

    # 保存到字节流
    output = io.BytesIO()
    new_wb.save(output)
    output.seek(0)
    return output.getvalue()


def archive_delete(person_id: int) -> bool:
    engine = get_engine()
    with engine.connect() as conn:
        exists = conn.execute(text("SELECT 1 FROM person WHERE id = :id"), {"id": person_id}).first()
    if not exists:
        return False
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM person WHERE id = :id"), {"id": person_id})
    return True


def status_counts() -> dict:
    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT status, COUNT(*) as count FROM person GROUP BY status")).mappings().all()
    return {"list": [{"status": r["status"], "count": int(r["count"])} for r in rows]}


def status_batch(ids: List[int], status: str) -> None:
    engine = get_engine()
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE person SET status = :s, updated_at = CURRENT_TIMESTAMP WHERE id IN :ids").bindparams(
                bindparam("ids", expanding=True)
            ),
            {"s": status, "ids": list(ids)},
        )
        on_site = 1 if status == "已进场" else 0 if status == "已离场" else None
        if on_site is not None:
            conn.execute(
                text("UPDATE person SET on_site = :v WHERE id IN :ids").bindparams(bindparam("ids", expanding=True)),
                {"v": on_site, "ids": list(ids)},
            )


def job_titles(*, flat: bool = False) -> dict:
    """
    返回工种列表。优先从 job_title_config 获取（支持层级），若无配置则从 person 表 DISTINCT。
    flat=1 时返回扁平列表供选择器使用。
    """
    cache_key = f"job_titles:flat={1 if flat else 0}"
    cached = _job_titles_cache_get(cache_key)
    if cached is not None:
        return cached
    engine = get_engine()
    try:
        with engine.connect() as conn:
            config_rows = conn.execute(
                text("SELECT id, code, name, parent_id, sort FROM job_title_config ORDER BY sort, id")
            ).mappings().all()
            if config_rows:
                # 构建树形结构
                by_id: Dict[int, Dict[str, Any]] = {int(r["id"]): dict(r) for r in config_rows}
                for r in by_id.values():
                    r["children"] = []
                roots: List[Dict[str, Any]] = []
                flat_names: List[str] = []
                for r in config_rows:
                    rid = int(r["id"])
                    parent_id = r.get("parent_id")
                    node = {**by_id[rid], "children": []}
                    flat_names.append(str(r["name"]))
                    if parent_id is None:
                        roots.append(node)
                    else:
                        pid = int(parent_id)
                        if pid in by_id and "children" in by_id[pid]:
                            by_id[pid]["children"].append(node)
                        else:
                            roots.append(node)
                # 递归收集叶子节点名称到 flat（用于选择）
                def collect_leaves(nodes: List[Dict]) -> List[str]:
                    out: List[str] = []
                    for n in nodes:
                        kids = n.get("children") or []
                        if not kids:
                            out.append(str(n.get("name", "")))
                        else:
                            out.extend(collect_leaves(kids))
                    return out

                flat_list = collect_leaves(roots) if roots else flat_names
                data = {"list": roots, "flat": flat_list}
                _job_titles_cache_set(cache_key, data)
                return data
    except Exception:  # noqa: BLE001
        pass

    # 回退：从 person 表获取
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT DISTINCT job_title FROM person WHERE TRIM(COALESCE(job_title,'')) != '' ORDER BY job_title")
        ).mappings().all()
    flat_list = [r["job_title"] for r in rows]
    data = {"list": flat_list, "flat": flat_list}
    _job_titles_cache_set(cache_key, data)
    return data


def face_verify_mark_passed(person_id: Optional[int]) -> None:
    if not person_id:
        return
    engine = get_engine()
    with engine.begin() as conn:
        conn.execute(
            text(
                "UPDATE person SET face_verified = 1, face_verified_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = :id"
            ),
            {"id": int(person_id)},
        )


def face_verify_status(person_id: int) -> Optional[dict]:
    engine = get_engine()
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT id, name, face_verified, face_verified_at FROM person WHERE id = :id"),
            {"id": person_id},
        ).mappings().first()
    if not row:
        return None
    return {
        "person_id": int(row["id"]),
        "name": row["name"],
        "face_verified": bool(row["face_verified"]),
        "face_verified_at": row["face_verified_at"],
    }


def person_simple_get(person_id: int) -> Optional[dict]:
    engine = get_engine()
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM person WHERE id = :id"), {"id": person_id}).mappings().first()
    return dict(row) if row else None


def certificates_list(person_id: int) -> Optional[dict]:
    engine = get_engine()
    with engine.connect() as conn:
        exists = conn.execute(text("SELECT 1 FROM person WHERE id = :id"), {"id": person_id}).first()
        if not exists:
            return None
        rows = conn.execute(
            text(
                """
                SELECT id, name, certificate_no, issue_date, expiry_date, status
                FROM person_certificate
                WHERE person_id = :id
                ORDER BY expiry_date DESC
                """
            ),
            {"id": person_id},
        ).mappings().all()
    return {"list": [dict(r) for r in rows]}


def certificate_create(person_id: int, body: Dict[str, Any]) -> Optional[int]:
    if not body.get("name"):
        raise ValueError("name 必填")
    engine = get_engine()
    with engine.begin() as conn:
        exists = conn.execute(text("SELECT 1 FROM person WHERE id = :id"), {"id": person_id}).first()
        if not exists:
            return None
        params = {
            "pid": person_id,
            "n": body["name"],
            "no": body.get("certificate_no"),
            "i": body.get("issue_date"),
            "e": body.get("expiry_date"),
            "s": body.get("status") or "valid",
        }
        if engine.dialect.name == "sqlite":
            rr = conn.execute(
                text(
                    """
                    INSERT INTO person_certificate (person_id, name, certificate_no, issue_date, expiry_date, status)
                    VALUES (:pid, :n, :no, :i, :e, :s)
                    """
                ),
                params,
            )
            return int(rr.lastrowid or 0)
        r = conn.execute(
            text(
                """
                INSERT INTO person_certificate (person_id, name, certificate_no, issue_date, expiry_date, status)
                VALUES (:pid, :n, :no, :i, :e, :s)
                RETURNING id
                """
            ),
            params,
        ).mappings().first()
        return int(r["id"])


def certificate_update(person_id: int, cert_id: int, patch: Dict[str, Any]) -> str:
    engine = get_engine()
    with engine.connect() as conn:
        pe = conn.execute(text("SELECT 1 FROM person WHERE id = :id"), {"id": person_id}).first()
        if not pe:
            return "no_person"
        ce = conn.execute(
            text("SELECT 1 FROM person_certificate WHERE id = :cid AND person_id = :pid"),
            {"cid": cert_id, "pid": person_id},
        ).first()
        if not ce:
            return "no_cert"

    updates = []
    params: Dict[str, Any] = {"cid": cert_id}
    for k in ("name", "certificate_no", "issue_date", "expiry_date", "status"):
        if k in patch:
            updates.append(f"{k} = :{k}")
            params[k] = patch.get(k)
    if not updates:
        raise ValueError("无有效字段")
    with engine.begin() as conn:
        set_clause = ", ".join(updates)
        query = text(f"UPDATE person_certificate SET {set_clause} WHERE id = :cid")
        conn.execute(query, params)
    return "ok"


def certificate_delete(person_id: int, cert_id: int) -> str:
    engine = get_engine()
    with engine.begin() as conn:
        ce = conn.execute(
            text("SELECT 1 FROM person_certificate WHERE id = :cid AND person_id = :pid"),
            {"cid": cert_id, "pid": person_id},
        ).first()
        if not ce:
            return "no_cert"
        conn.execute(text("DELETE FROM person_certificate WHERE id = :cid"), {"cid": cert_id})
    return "ok"

